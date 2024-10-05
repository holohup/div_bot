from aiogram import Bot, Dispatcher, F
from aiogram.types import FSInputFile
from aiogram.filters import Command
from aiogram.types import Message
import openpyxl
from openpyxl.styles import Border, Side
from settings import STORAGE
from users import IsAdmin, UserHandler, IsApproved
from service import DividendCounter, DISCOUNT_RATE, IndexCounter
from zoneinfo import ZoneInfo
from datetime import datetime
from settings import TG_BOT_TOKEN, TG_ADMIN_IDS
import pandas as pd
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
bot = Bot(TG_BOT_TOKEN)
dp = Dispatcher()
user_handler = UserHandler(STORAGE)
moscow_tz = ZoneInfo('Europe/Moscow')


def parse_command(cmd: str) -> str:
    return cmd.split()[-1]


@dp.message(IsAdmin() and Command(commands='approve'))
async def approve_user(message: Message):
    id_ = int(parse_command(message.text))
    await message.answer(text=user_handler.approve_user(id_))


@dp.message(Command(commands='start'))
async def welcome_new_user(message: Message):
    id_ = message.from_user.id
    name = message.from_user.full_name
    user_handler.register_user(id_)
    await message.answer('Добро пожаловать в мир дивибота!')
    await bot.send_message(
        chat_id=TG_ADMIN_IDS,
        text=f'У нас новый пользователь {name} id следующим сообщением'
    )
    await bot.send_message(
        chat_id=TG_ADMIN_IDS,
        text=f'{id_}'
    )


@dp.message(Command(commands='list'))
async def process_ticker_list(message: Message):
    await message.answer('Список тикеров, для которых есть хотя бы один фьюч:')
    await message.answer(
        await DividendCounter(STORAGE).list_available_tickers()
    )


@dp.message(IsApproved(), Command(commands='d'))
async def process_details(message: Message):
    ticker = parse_command(message.text)
    moscow_time = datetime.now(moscow_tz)
    formatted_time = moscow_time.strftime('%H:%M:%S %d-%m-%y')
    futures, stock = await DividendCounter(STORAGE, ticker).count()
    df_string = format_details_message(futures)
    await message.reply(
        f"Futures for {stock.iloc[0].ticker} ({formatted_time}):\n"
        f"<pre>{df_string}</pre>",
        parse_mode='HTML'
    )


@dp.message(IsApproved(), F.text.lower() == 'ind')
async def process_index(message: Message):
    result = await IndexCounter().run()
    await message.reply(
        f'Current index futures {result}'
    )


@dp.message(IsApproved(), Command(commands='all'))
async def process_full_list(message: Message):
    filename = await DividendCounter(STORAGE).count_all()
    
    # Load the existing workbook
    existing_wb = openpyxl.load_workbook(filename)
    existing_ws = existing_wb.active
    
    # Create a new workbook for the final output
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    
    # Add meta information
    meta_info = [
        f"Репорт по дивам {datetime.now(moscow_tz).strftime('%d-%m-%Y %H:%M:%S')}",
        f'Ставка: {DISCOUNT_RATE}% годовых'
    ]
    
    for index, info in enumerate(meta_info, start=1):
        new_ws.merge_cells(start_row=index, start_column=1, end_row=index, end_column=existing_ws.max_column)
        new_ws.cell(row=index, column=1).value = info
    
    # Start appending data from the existing file after the metadata
    start_row = len(meta_info) + 2  # One additional row for spacing
    
    for row in existing_ws.iter_rows(min_row=1, max_row=existing_ws.max_row, max_col=existing_ws.max_column):
        for col_index, cell in enumerate(row, start=1):
            new_ws.cell(row=start_row, column=col_index).value = cell.value
        start_row += 1
    
    # Apply borders to the new worksheet as needed
    thin_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    current_ticker = None

    # Iterate through the rows in the new sheet to apply borders
    for row in range(len(meta_info) + 2, new_ws.max_row + 1):
        ticker = new_ws.cell(row=row, column=2).value
        if ticker != current_ticker:
            if current_ticker is not None:
                for c in range(1, new_ws.max_column + 1):
                    new_ws.cell(row=row, column=c).border = Border(top=thin_border.top)

            current_ticker = ticker

    # Save the new workbook
    new_filename = 'final_report.xlsx'
    new_wb.save(new_filename)

    # Send the new file to the user
    result = FSInputFile(new_filename)
    await message.answer_document(result)


@dp.message(IsApproved())
async def process_other_answers(message: Message):
    moscow_time = datetime.now(moscow_tz)
    formatted_time = moscow_time.strftime('%H:%M:%S %d-%m-%y')
    try:
        futures, stock = await DividendCounter(STORAGE, message.text).count()
        df_string = format_message(futures)
        await message.reply(
            f"Futures for {stock.iloc[0].ticker} ({formatted_time}), discount rate = {DISCOUNT_RATE}:\n"
            f"<pre>{df_string}</pre>",
            parse_mode='HTML'
        )

    except Exception as e:
        await message.answer(str(e))


def format_message(futures):
    short_columns = {
            'expiration_date': 'expires',
            'div_percent': 'div%',
            'dividend': 'div'
        }
    futures = futures.rename(columns=short_columns)
    futures['div'] = futures['div'].round(2)
    futures['div%'] = futures['div%'].round(2)
    futures['expires'] = pd.to_datetime(futures['expires'])
    futures['expires'] = futures['expires'].dt.strftime('%d.%m.%y')
    df_string = futures[
            ['ticker', 'expires', 'days', 'div', 'div%']
        ].to_string(index=False)

    return df_string


def format_details_message(futures):
    short_columns = {
            'expiration_date': 'expires',
            'div_percent': 'div%',
            'dividend': 'div',
            'sell_margin': 's_m',
            'buy_margin': 'b_m'
        }
    futures = futures.rename(columns=short_columns)
    futures['div'] = futures['div'].round(2)
    futures['div%'] = futures['div%'].round(2)
    futures['expires'] = pd.to_datetime(futures['expires'])
    futures['expires'] = futures['expires'].dt.strftime('%d.%m.%y')
    futures['fair'] = futures['fair'].round(2)
    futures['current'] = futures['current'].round(2)
    print(futures)
    df_string = futures[
            ['ticker', 'expires', 'days', 'div', 'div%', 'current', 'fair', 's_m', 'b_m']
        ].to_string(index=False)

    return df_string

if __name__ == '__main__':
    try:
        dp.run_polling(bot)
    except Exception as e:
        logger.exception(str(e.__traceback__))
